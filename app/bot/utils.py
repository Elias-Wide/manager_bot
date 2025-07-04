import os

import random
import string
from io import BytesIO
from aiogram.types import ContentType, Message
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.core.constants import (
    FMT_JPG,
)
from app.bot.init_bot import bot


async def generate_filename() -> str:
    """
    Generate a random filename consisting of 10 characters (letters and digits).

    Returns:
        str: The generated filename.
    """
    filename = [
        random.choice(
            string.ascii_lowercase + string.digits if i != 5 else string.ascii_uppercase
        )
        for i in range(10)
    ]
    return "".join(filename)


async def is_file_in_dir(name, path) -> None:
    """
    Check if a file with the given name exists in the specified directory.

    Args:
        name (str): The filename to search for.
        path (str): The directory path.

    Returns:
        bool: True if the file exists, otherwise False.
    """
    for root, dirs, files in os.walk(path):
        if name in files:
            return True
        return False


async def delete_file(path: str, file_name: str):
    """
    Delete a file with the specified name in the given directory.

    Args:
        path (str): The directory path.
        file_name (str): The name of the file to delete.
    """
    try:
        os.remove(path / (file_name + FMT_JPG))
    except:
        print("Error deleting file")


async def delete_files_in_folder(folder_path: str):
    """
    Delete all files in the specified folder.

    Args:
        folder_path (str): The path to the folder.
    """
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        try:
            if os.path.isfile(file_path):
                os.remove(file_path)
        except Exception as e:
            print(f"Error deleting file {file_path}. {e}")


async def download_file(file, destination) -> str:
    """
    Download a file from the bot and save it to the specified directory.

    Args:
        file: The file object from Telegram.
        destination: The directory to save the file.

    Returns:
        str: The generated filename (without extension).
    """
    file_name = await generate_filename()
    filename_with_format = file_name + FMT_JPG
    path = destination / (file_name + FMT_JPG)
    file_from_bot = await bot.get_file(file.file_id)
    destination_file = await bot.download_file(
        file_from_bot.file_path, os.path.join(os.getcwd(), path)
    )
    return file_name


async def delete_file(path: str):
    """
    Delete a file at the specified path.

    Args:
        path (str): The full path to the file.
    """
    try:
        os.remove(path)
    except:
        print("Error deleting file")


async def read_excel_file(message: Message):
    """
    Process data from an Excel file with office data.

    Args:
        message (Message): The incoming message containing the Excel file.

    Returns:
        list[dict]: List of dictionaries with office data.
    """

    xlsx_file_in_buffer = await download_file_from_bot(message)
    workbook = openpyxl.load_workbook(xlsx_file_in_buffer)
    sheet = workbook.active
    return [
        {
            "id": row[0],
            "addres": row[1],
            "name": row[3],
        }
        for row in sheet.iter_rows(values_only=True)
    ]


async def download_file_from_bot(message: Message) -> BytesIO:
    """
    Download a file from the bot into a memory buffer.

    Args:
        message (Message): The incoming message containing the file.

    Returns:
        BytesIO: The buffer with the downloaded file.
    """

    buffer = BytesIO()
    if message.content_type == ContentType.PHOTO:
        file_from_bot = await bot.get_file(message.photo[-1].file_id)
    elif message.content_type == ContentType.DOCUMENT:
        file_from_bot = await bot.get_file(message.document.file_id)
    return await bot.download_file(file_from_bot.file_path, buffer)


async def create_excel_report(region_report_data: list[tuple]) -> BytesIO:
    """
    Generate an Excel report in memory from a list of tuples with office report data.

    This function creates an Excel file with the following logic:
    - Initializes a new workbook and sets the active worksheet's title to "Отчеты по офисам".
    - Adds a header row with columns: "Пункт", "ID", "Менеджер", "Отчет прихода".
    - Applies bold font and centered alignment to the header cells.
    - For each row in region_report_data:
        - If the "Менеджер" field (index 2) is a list or tuple, joins its elements with a newline character so each manager appears on a new line in the cell.
        - Appends the row to the worksheet.
        - If any cell in the row contains the value "нет отчета", the entire row is filled with a red background; otherwise, it is filled with a green background.
        - The "Менеджер" column is set to wrap text and center alignment to properly display multiple managers.
        - A thin bottom border is added to each cell in the row, except for the last row.
    - Sets custom column widths for better readability.
    - Saves the workbook to a BytesIO buffer and returns it, ready for sending as a file.

    Args:
        region_report_data (list[tuple]): List of tuples, each representing a row of office report data.

    Returns:
        BytesIO: The buffer containing the generated Excel file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчеты по офисам"

    headers = ["Пункт", "ID", "Менеджер", "Отчет прихода"]
    ws.append(headers)
    header_font = Font(bold=True)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )

    thin_border = Border(bottom=Side(style="thin", color="000000"))

    for row_idx, row in enumerate(region_report_data, start=2):
        row = list(row)
        print(row)
        if isinstance(row[2], (list, tuple)):
            row[2] = "\n".join(str(m) for m in row[2])
        if row[3] == False:
            row[3] = "НЕТ ОТЧЕТА"
        ws.append(row)
        fill = red_fill if any(cell == "НЕТ ОТЧЕТА" for cell in row) else green_fill
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            if col_idx == 3:
                cell.alignment = Alignment(
                    wrap_text=True, horizontal="center", vertical="center"
                )
        if row_idx < len(region_report_data) + 1:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).border = thin_border

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
